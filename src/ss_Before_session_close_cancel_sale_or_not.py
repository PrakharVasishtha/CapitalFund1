"""
ss_Before_session_close_cancel_sale_or_not.py
===============================================
Monitors pre-open / indicative prices of IPO stocks on listing day (scheduled at 09:32 AM).

Checks open sell orders placed by ss_sale_order_on_lc_on_start_of_ss (special_session_status == 1).
If loss/discount thresholds are exceeded:
  - category == "sme" and loss_percent > 0
  - OR category == "mb" and loss_percent > 11.9
  where loss_percent = ((issue_price - indicative_price) / issue_price) * 100.0

Action taken per README.md:
  - IF Cancel triggered:
      Cancels sell order via zerodha_cancel_order()
      Updates special_session_status = 3 (Not sold in special session)
      Updates regular_session_status = 1 (Regular session sell started if not sold in special session)
  - IF Cancel NOT triggered (LC sell order remains active & executes):
      Updates special_session_status = 2 (Sold in special pre-open session)
      Updates regular_session_status = 0 (Not started)
"""
import sys
import os
import time
import openpyxl
from datetime import datetime

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from Base import load_credentials
from allotment_update import get_allotted_holdings_path
from special_session_indicative_price_nse import get_ipo_indicative_price
from special_sesion_zerodha_sell import zerodha_cancel_order

CREDENTIALS_FILE = "credentials.json"


def sale_order_cancel_or_not():
    """
    Monitors IPO indicative prices during pre-open session (scheduled at 09:32 AM).
    Evaluates pre-open sell orders (special_session_status == 1).
    Updates special_session_status and regular_session_status in allotted_holdings.xlsx
    according to action taken based on README.md status code definitions.
    """
    print("-----------special_session_monitor (09:32 AM Cancel Check)---------")

    users = load_credentials(CREDENTIALS_FILE)
    if not users:
        print("special_session_monitor: No user credentials found.")
        return

    excel_path = get_allotted_holdings_path()
    if not os.path.exists(excel_path):
        print(f"special_session_monitor: Excel file not found at '{excel_path}'")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"special_session_monitor: Error loading workbook '{excel_path}': {e}")
        return

    file_changed = False

    for user in users:
        uci_user = str(user.get("uci"))
        client_id = user.get("broker_client_id")
        password_user = user.get("password_broker")
        totp_broker = user.get("topt_broker") or user.get("totp_broker")

        if uci_user not in wb.sheetnames:
            continue

        ws = wb[uci_user]

        for r in range(2, ws.max_row + 1):
            security_symbol = ws.cell(r, 1).value
            if not security_symbol or str(security_symbol).strip().lower() in ["none", ""]:
                continue

            security_symbol = str(security_symbol).strip()
            lot_size = ws.cell(r, 2).value or 1
            try:
                lot_size = int(lot_size)
            except (ValueError, TypeError):
                lot_size = 1

            issue_price = ws.cell(r, 3).value or 0.0
            try:
                issue_price = float(issue_price)
            except (ValueError, TypeError):
                issue_price = 0.0

            exchange = str(ws.cell(r, 6).value or "NSE").strip().upper()

            # Identify column positions (Col 8 = special_session_status, Col 11 = regular_session_status)
            spl_status_col = 8 if ws.cell(r, 8).value is not None or ws.cell(1, 8).value == "special_session_status" else 7
            reg_status_col = 11 if ws.cell(r, 11).value is not None or ws.cell(1, 11).value == "regular_session_status" else 10

            status_cell = ws.cell(r, spl_status_col).value
            try:
                spl_status = int(status_cell) if status_cell is not None else 0
            except (ValueError, TypeError):
                spl_status = 0

            # Only check orders placed at start of special session (special_session_status == 1)
            if spl_status != 1:
                continue

            category = "sme" if lot_size >= 100 else "mb"

            print(f"\nChecking [{uci_user}] Symbol: '{security_symbol}' | Exchange: {exchange} | Issue Price: {issue_price} | Category: {category}")

            # Fetch pre-open indicative price (IEP)
            try:
                price_info = get_ipo_indicative_price(symbol=security_symbol, exchange=exchange)
            except Exception as e:
                print(f"Error fetching indicative price for {security_symbol}: {e}")
                price_info = {"indicative_price": 0.0}

            indicative_price = price_info.get("indicative_price", 0.0) if isinstance(price_info, dict) else float(price_info or 0.0)
            print(f"Result for {security_symbol}: IEP = {indicative_price}")

            loss_percent = 0.0
            if issue_price > 0 and indicative_price > 0 and issue_price > indicative_price:
                loss_percent = ((issue_price - indicative_price) / issue_price) * 100.0

            print(f"Calculated discount/loss %: {loss_percent:.2f}%")

            # Decision rules to CANCEL order per README:
            # SME loss_percent > 0  OR  Mainboard loss_percent > 11.9
            should_cancel = False
            if category == "sme" and loss_percent > 0:
                should_cancel = True
            elif category == "mb" and loss_percent > 11.9:
                should_cancel = True

            if should_cancel:
                print(f"Triggering order CANCEL for {security_symbol} (User: {uci_user}, Loss %: {loss_percent:.2f}%)...")
                if client_id and password_user and totp_broker:
                    try:
                        success, msg = zerodha_cancel_order(
                            user_id=client_id,
                            password=password_user,
                            totp_secret=totp_broker,
                            security_symbol=security_symbol
                        )
                        print(f"Cancel execution output: {msg}")
                        if success:
                            # Per README.md:
                            # special_session_status = 3 ("Not sold in special session")
                            # regular_session_status = 1 ("Regular session sell started if not sold in special session")
                            ws.cell(r, spl_status_col, 3)
                            ws.cell(r, reg_status_col, 1)
                            file_changed = True
                    except Exception as e:
                        print(f"Failed to cancel order for {security_symbol}: {e}")
                else:
                    print(f"Missing credentials for {uci_user}, cannot cancel sell order.")
            else:
                # Order NOT canceled -> allowed to execute in special pre-open session
                # Per README.md:
                # special_session_status = 2 ("Sold in special pre-open session")
                # regular_session_status = 0 ("Not started")
                print(f"Keeping LC sell order active for {security_symbol} (Sold in special session).")
                ws.cell(r, spl_status_col, 2)
                ws.cell(r, reg_status_col, 0)
                file_changed = True

    if file_changed:
        try:
            wb.save(excel_path)
            print("\nspecial_session_monitor: Successfully saved status updates to allotted_holdings.xlsx per README.md")
        except Exception as e:
            print(f"special_session_monitor: Error saving workbook: {e}")

    wb.close()
    print("special_session_monitor pass completed.")


if __name__ == "__main__":
    sale_order_cancel_or_not()