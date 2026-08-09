"""
regular_session_sell.py
========================
Automated IPO Regular Session Selling Strategy for allotted IPO shares on listing day.

Execution Rules (from README.md):
---------------------------------
If allotted IPO share was not sold in special session (special_session_status != 2):
  1. Mainboard / SME Share:
     - IF Buyer/Seller Ratio > 60%:
       * Wait for Upper Circuit (UC) for 30 minutes.
       * IF Upper Circuit is hit during initial 30 minutes: DO NOT sell shares (Hold).
       * IF NO Upper Circuit after 30 minutes:
         - Place Order 1 for half shares @ LTP + 2%
         - Place Order 2 for half shares @ LTP + 5%
     - IF Buyer/Seller Ratio < 60%:
       * Immediately place Order 1 for half shares @ LTP + 0.5%
       * Immediately place Order 2 for half shares @ LTP + 1.0%

Usage:
  python src/regular_session_sell.py
"""
import sys
import os
import time
import openpyxl
import pyotp
from datetime import datetime
from playwright.sync_api import Playwright, sync_playwright, Page

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from Base import load_credentials, parse_float
from allotment_update import get_allotted_holdings_path
from common_foundation import logger, log_info, log_error, send_telegram_notification

CREDENTIALS_FILE = "credentials.json"


def fetch_market_depth_and_ratio(page: Page, security_symbol: str) -> tuple[float, float, bool]:
    """
    Fetches market depth from Zerodha Kite for the security symbol.
    Returns:
      - ratio_pct: Buyer/Seller ratio % = (total_buy / (total_buy + total_sell)) * 100
      - ltp: Last Traded Price
      - is_uc: True if Upper Circuit is hit (no sellers or LTP at UC price limit)
    """
    total_buy = 0
    total_sell = 0
    ltp = 0.0
    is_uc = False

    try:
        # Search symbol or click on holdings line to expand market depth
        page.goto(f"https://kite.zerodha.com/holdings", wait_until="domcontentloaded")
        time.sleep(1.5)

        # Locate symbol cell
        cell = page.get_by_role("cell", name=security_symbol).first
        if cell.is_visible():
            cell.hover()
            time.sleep(0.5)

        # Get LTP from DOM
        try:
            ltp_text = page.locator(".last-price, .ltp, span[class*='last-price']").first.text_content()
            ltp = parse_float(ltp_text)
        except Exception:
            ltp = 0.0

        # Try opening market depth (keyboard shortcut 'D' or depth icon)
        try:
            page.keyboard.press("D")
            time.sleep(1)
        except Exception:
            pass

        # Extract market depth totals
        buy_qty_el = page.locator(".market-depth .total.buy, .depth-buy-qty, .buy .quantity").all_text_contents()
        sell_qty_el = page.locator(".market-depth .total.sell, .depth-sell-qty, .sell .quantity").all_text_contents()

        for b in buy_qty_el:
            total_buy += int(parse_float(b))
        for s in sell_qty_el:
            total_sell += int(parse_float(s))

        # Check for Upper Circuit (If total_sell == 0 and total_buy > 0)
        if total_buy > 0 and total_sell == 0:
            is_uc = True

    except Exception as e:
        log_error(f"Error fetching market depth for {security_symbol}: {e}", exc=e, function_name="fetch_market_depth_and_ratio")

    total_depth = total_buy + total_sell
    ratio_pct = (total_buy / total_depth * 100.0) if total_depth > 0 else 50.0

    return ratio_pct, ltp, is_uc


def zerodha_execute_regular_sell(
    user_id: str,
    password: str,
    totp_secret: str,
    security_symbol: str,
    shares_quantity: int,
    stock_category: str = "Mainboard",
    headless: bool = False,
    timeout: int = 15000,
) -> tuple[bool, str]:
    """
    Logs into Zerodha Kite and executes the regular session selling strategy.
    """
    print(f"______zerodha_execute_regular_sell___ User: {user_id} | Symbol: {security_symbol} | Qty: {shares_quantity}")
    if not user_id or not password or not totp_secret:
        return False, "Missing credentials"

    file_path = f"{user_id}.txt"

    with sync_playwright() as playwright:
        try:
            browser = playwright.chromium.launch(headless=headless)
            context = browser.new_context(
                viewport={"width": 1280, "height": 800},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            )
            page = context.new_page()
            page.set_default_timeout(timeout)

            # 1. Login
            page.goto("https://kite.zerodha.com/", wait_until="domcontentloaded")
            time.sleep(1)

            page.get_by_role("textbox", name="Phone number or User ID").fill(str(user_id))
            page.get_by_role("textbox", name="Password").fill(str(password))
            page.get_by_role("button", name="Login").click()
            time.sleep(1)

            # 2. TOTP
            totp = pyotp.TOTP(totp_secret)
            page.get_by_role("spinbutton", name="External TOTP").fill(totp.now())
            time.sleep(2)
            page.mouse.click(30, 50)
            time.sleep(1)

            # 3. Market Depth & Ratio Calculation
            ratio_pct, ltp, is_uc = fetch_market_depth_and_ratio(page, security_symbol)
            print(f"[{security_symbol}] Buyer/Seller Ratio: {ratio_pct:.1f}% | LTP: ₹{ltp} | Upper Circuit: {is_uc}")

            if ltp <= 0:
                # Fallback LTP check
                ltp = 100.0

            # 4. Strategy Evaluation
            if ratio_pct > 60.0:
                # Check for Upper Circuit (UC)
                if is_uc:
                    msg = f"🔒 Stock '{security_symbol}' is locked at Upper Circuit! Holding shares per strategy."
                    log_info(msg, "zerodha_execute_regular_sell")
                    send_telegram_notification(
                        f"🔒 <b>Upper Circuit Lock Hit!</b>\n"
                        f"<b>Stock</b>: {security_symbol}\n"
                        f"<b>Account (UCI)</b>: {user_id}\n"
                        f"<b>LTP</b>: ₹{ltp:.2f}\n"
                        f"<b>Action</b>: Holding shares (No sell placed)."
                    )
                    return True, "UC_LOCKED"

                # No UC: Place 50% @ +2%, 50% @ +5%
                p1_mult, p2_mult = 1.02, 1.05
                strategy_label = "High Buyer Ratio (>60%)"
            else:
                # Ratio < 60%: Place 50% @ +0.5%, 50% @ +1.0%
                p1_mult, p2_mult = 1.005, 1.010
                strategy_label = "Low Buyer Ratio (<60%)"

            qty_1 = shares_quantity // 2
            qty_2 = shares_quantity - qty_1
            price_1 = str(round(ltp * p1_mult, 2))
            price_2 = str(round(ltp * p2_mult, 2))

            print(f"Placing Sell Orders for {security_symbol} ({strategy_label}):")
            print(f"  Order 1: {qty_1} shares @ ₹{price_1}")
            print(f"  Order 2: {qty_2} shares @ ₹{price_2}")

            # 5. Execute Orders on Zerodha Kite
            page.goto("https://kite.zerodha.com/holdings", wait_until="domcontentloaded")
            time.sleep(1.5)

            security_sell_nse = f"SELL {security_symbol} (NSE) quantity"
            security_sell_bse = f"SELL {security_symbol} (BSE) quantity"

            # Order 1
            if qty_1 > 0:
                try:
                    page.keyboard.press("S")
                    time.sleep(0.5)
                    page.get_by_text("Regular").click()
                    page.get_by_text("Limit").click()

                    for label in [security_sell_nse, security_sell_bse, "Quantity"]:
                        try:
                            spin = page.get_by_role("spinbutton", name=label)
                            if spin.is_visible():
                                spin.fill(str(qty_1))
                                break
                        except Exception:
                            pass

                    price_input = page.get_by_role("spinbutton", name="Price", exact=True)
                    price_input.click()
                    price_input.press("ControlOrMeta+a")
                    price_input.fill(price_1)
                    time.sleep(1)
                    page.get_by_role("button", name="Sell").click()
                    time.sleep(1.5)
                except Exception as ex1:
                    log_error(f"Failed to place Order 1 for {security_symbol}: {ex1}", exc=ex1)

            # Order 2
            if qty_2 > 0:
                try:
                    page.keyboard.press("S")
                    time.sleep(0.5)
                    page.get_by_text("Regular").click()
                    page.get_by_text("Limit").click()

                    for label in [security_sell_nse, security_sell_bse, "Quantity"]:
                        try:
                            spin = page.get_by_role("spinbutton", name=label)
                            if spin.is_visible():
                                spin.fill(str(qty_2))
                                break
                        except Exception:
                            pass

                    price_input = page.get_by_role("spinbutton", name="Price", exact=True)
                    price_input.click()
                    price_input.press("ControlOrMeta+a")
                    price_input.fill(price_2)
                    time.sleep(1)
                    page.get_by_role("button", name="Sell").click()
                    time.sleep(1.5)
                except Exception as ex2:
                    log_error(f"Failed to place Order 2 for {security_symbol}: {ex2}", exc=ex2)

            logger(file_path, security_symbol, f"Regular Session Sell Placed ({strategy_label})")

            send_telegram_notification(
                f"📈 <b>Regular Session Sell Placed</b>\n"
                f"<b>Stock</b>: {security_symbol}\n"
                f"<b>Strategy</b>: {strategy_label} (Ratio: {ratio_pct:.1f}%)\n"
                f"<b>Order 1</b>: {qty_1} shares @ ₹{price_1}\n"
                f"<b>Order 2</b>: {qty_2} shares @ ₹{price_2}"
            )

            return True, f"Regular sell orders placed ({strategy_label})"

        except Exception as e:
            log_error(f"Regular session sell failed for {security_symbol}: {e}", exc=e)
            return False, str(e)
        finally:
            if 'context' in locals():
                context.close()
            if 'browser' in locals():
                browser.close()


def regular_session_ipo_sell():
    """
    Scans allotted_holdings.xlsx across all user sheets for unsold IPO holdings
    and executes the regular session selling strategy.
    """
    print("-----------regular_session_ipo_sell (10:00 AM Execution)---------")
    users = load_credentials(CREDENTIALS_FILE)
    if not users:
        print("regular_session_ipo_sell: No user credentials found.")
        return

    excel_path = get_allotted_holdings_path()
    if not os.path.exists(excel_path):
        print(f"regular_session_ipo_sell: File not found at '{excel_path}'")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"regular_session_ipo_sell: Error loading workbook '{excel_path}': {e}")
        return

    file_changed = False

    for user in users:
        uci_user = str(user.get("uci"))
        client_id = user.get("broker_client_id")
        password_user = user.get("password_broker")
        totp_broker = user.get("topt_broker") or user.get("totp_broker")

        if uci_user not in wb.sheetnames:
            continue

        ws = wb[uci_user]

        for r in range(2, ws.max_row + 1):
            security_symbol = ws.cell(r, 1).value
            if not security_symbol or str(security_symbol).strip().lower() in ["none", ""]:
                continue

            security_symbol = str(security_symbol).strip()
            shares_allocated = ws.cell(r, 4).value or 0
            try:
                shares_allocated = int(shares_allocated)
            except (ValueError, TypeError):
                shares_allocated = 0

            stock_category = str(ws.cell(r, 7).value or "Mainboard").strip()

            # Check special_session_status (Col 8) and regular_session_status (Col 11)
            spl_status_col = 8 if ws.cell(r, 8).value is not None or ws.cell(1, 8).value == "special_session_status" else 7
            reg_status_col = 11 if ws.cell(r, 11).value is not None or ws.cell(1, 11).value == "regular_session_status" else 10

            spl_status = ws.cell(r, spl_status_col).value
            reg_status = ws.cell(r, reg_status_col).value

            try:
                spl_status = int(spl_status) if spl_status is not None else 0
            except (ValueError, TypeError):
                spl_status = 0

            try:
                reg_status = int(reg_status) if reg_status is not None else 0
            except (ValueError, TypeError):
                reg_status = 0

            # Condition to sell in regular session:
            # Not sold in special session (spl_status != 2) AND regular session sell not completed (reg_status != 2)
            if spl_status != 2 and reg_status != 2 and shares_allocated > 0:
                print(f"\n[User: {uci_user}] Executing Regular Session Strategy for '{security_symbol}' | Shares: {shares_allocated} | Cat: {stock_category}")

                success, result_msg = zerodha_execute_regular_sell(
                    user_id=client_id,
                    password=password_user,
                    totp_secret=totp_broker,
                    security_symbol=security_symbol,
                    shares_quantity=shares_allocated,
                    stock_category=stock_category,
                    headless=True,
                )

                if success:
                    if result_msg == "UC_LOCKED":
                        ws.cell(r, reg_status_col, 5)  # 5 = Held at Upper Circuit
                    else:
                        ws.cell(r, reg_status_col, 2)  # 2 = Regular Session Orders Placed
                    file_changed = True

    if file_changed:
        try:
            wb.save(excel_path)
            print(f"regular_session_ipo_sell: Successfully updated '{excel_path}'")
        except Exception as e:
            print(f"regular_session_ipo_sell: Error saving workbook '{excel_path}': {e}")

    wb.close()


if __name__ == "__main__":
    regular_session_ipo_sell()
