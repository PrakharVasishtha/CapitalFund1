from datetime import date

import openpyxl

from src import Base
from src.kotak_base import apply_to_ipo


def IPO_to_apply():
    print("--------IPO to Apply-------")
    IPO_sme_1 = []
    IPO_sme_2 = []
    IPO_mb_1 = []
    IPO_mb_2 = []

    row_sme = Base.get_last_row_sme() - 1
    row_mb = Base.get_last_row_mb() - 1


    path = '../General.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    sme_ws = wb['IPOSME']
    main_ws = wb['IPOMB']

    for i in range(0, 9):
        rw = row_sme - i
        apply = sme_ws.cell(rw, 42).value
        name = sme_ws.cell(rw, 2).value
        name = name[:15]
        if apply == 1:
            close_date = sme_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_sme_1.append(name)

        if apply == 2:
            close_date = sme_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_sme_2.append(name)

    for i in range(0, 9):
        rw = row_mb - i
        apply = main_ws.cell(rw, 42).value
        name = main_ws.cell(rw, 2).value
        name = name[:15]
        if apply == 1:
            close_date = main_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_mb_1.append(name)

        if apply == 2:
            close_date = main_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_mb_2.append(name)

    return IPO_sme_1, IPO_sme_2, IPO_mb_1, IPO_mb_2

def apply_ipo_sme(ipo):
    print("Appling IPO:", ipo)

def apply_ipo_mb(ipo):
    print("Appling IPO:", ipo)
    #mb first try in snii category if not sufficient funds, it will try in retail.

def ipo_application():
    print("--------IPO Application-------")
    all=IPO_to_apply()
    IPO_sme_1 = all[0]
    IPO_sme_2 = all[1]
    IPO_mb_1 = all[2]
    IPO_mb_2 = all[3]

    for ipo in IPO_mb_2:
        print(ipo)
        apply_to_ipo(
            ipo_name=ipo,
            bank_user="jhkh",
            bank_pwd = "hkhk",
            type = "MB",
            headless = False
        )

    for ipo in IPO_sme_2:
        print(ipo)
        apply_ipo_sme(ipo)

    for ipo in IPO_mb_1:
        print(ipo)
        apply_ipo_mb(ipo)

    for ipo in IPO_sme_1:
        print(ipo)
        apply_ipo_sme(ipo)

print(IPO_to_apply())
#ipo_application()