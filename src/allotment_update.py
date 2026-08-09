"""
allotment_update.py
===================
Updates allotted_holdings.xlsx with core stock details for newly detected allotments.

Key functions:
  get_allotted_holdings_path()     -- Resolves the absolute path to allotted_holdings.xlsx
                                      regardless of current working directory.
  get_stock_type(holding_symbol)   -- Returns 'Mainboard' or 'SME'
  get_exchange(holding_symbol)     -- Returns 'NSE', 'BSE', or 'NSE / BSE'
  get_issue_price(holding_symbol)  -- Returns issue price float
  get_lot_size(holding_symbol)     -- Returns lot size int
  excel_holdings(usr_id, holding_symbol)
    - If holding_symbol is provided and not already in the user's sheet, appends a new row
      with special_session_status = 5 (pending enrichment).
    - Iterates rows with special_session_status == 5 and populates ONLY:
        Col 1: security_name (stock symbol)
        Col 2: lot_size
        Col 3: issue_price
        Col 6: exchange (NSE / BSE)
        Col 7: stock_category (Mainboard / SME)
    - Saves updated data back to allotted_holdings.xlsx.

Column reference for allotted_holdings.xlsx (1-indexed, Row 1 = header):
  Col 1  : security_name
  Col 2  : lot_size
  Col 3  : issue_price
  Col 4  : shares_allocated
  Col 5  : lots_issued
  Col 6  : exchange (NSE / BSE)
  Col 7  : stock_category (Mainboard / SME)
  Col 8  : special_session_status (0=not started, 1=started, 2=sold, 3=unsold, 5=pending)
  Col 11 : regular_session_status (0=not started, 2=sold, 3=unsold, 4=transferred)
"""
import sys
import os

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from allotment_fetch import fetch_allotment_holdings
import openpyxl
import requests


def get_allotted_holdings_path() -> str:
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(base_dir, "allotted_holdings.xlsx")
    if os.path.exists(path):
        return path
    if os.path.exists("allotted_holdings.xlsx"):
        return "allotted_holdings.xlsx"
    return path


def get_stock_full_details(symbol_or_name: str, existing_lot_size: int = 1, existing_price: float = 0.0) -> tuple[str, str, float, int]:
    """
    Fetches stock_type ('Mainboard' or 'SME'), listing exchange ('NSE', 'BSE', or 'NSE / BSE'),
    issue_price (float), and lot_size (int) for a given stock symbol or IPO name.
    """
    if not symbol_or_name or str(symbol_or_name).strip() == "":
        return "Mainboard", "NSE", 0.0, 1

    symbol_clean = str(symbol_or_name).strip()
    stock_type = "Mainboard"
    exchange = "NSE"
    issue_price = existing_price if existing_price > 0 else 0.0
    lot_size = existing_lot_size if existing_lot_size > 1 else 1

    # 1. Check General.xlsx cache
    gen_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "General.xlsx")
    if not os.path.exists(gen_path) and os.path.exists("General.xlsx"):
        gen_path = "General.xlsx"

    chittorgarh_url = None
    if os.path.exists(gen_path):
        try:
            wb = openpyxl.load_workbook(gen_path, data_only=True)
            for sheet_name, cat in [("IPOSME", "SME"), ("IPOMB", "Mainboard")]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for r in range(2, ws.max_row + 1):
                        name_val = str(ws.cell(r, 2).value or "").strip()
                        url_val = str(ws.cell(r, 1).value or "").strip()
                        if symbol_clean.lower() in name_val.lower() or name_val.lower() in symbol_clean.lower():
                            stock_type = cat
                            if url_val.startswith("http"):
                                chittorgarh_url = url_val
                            break
                    if chittorgarh_url:
                        break
            wb.close()
        except Exception:
            pass

    # 2. Try Chittorgarh scraping
    urls_to_try = []
    if chittorgarh_url:
        urls_to_try.append(chittorgarh_url)
    slug = symbol_clean.lower().replace(" ", "-").replace("_", "-")
    urls_to_try.append(f"https://www.chittorgarh.com/ipo/{slug}-ipo/")

    try:
        import cloudscraper
        import bs4
        import re
        scraper = cloudscraper.create_scraper(browser={"browser": "chrome", "platform": "windows", "desktop": True})
        for u in urls_to_try:
            try:
                res = scraper.get(u, timeout=10)
                if res.status_code == 200:
                    soup = bs4.BeautifulSoup(res.text, "html.parser")
                    for tr in soup.find_all("tr"):
                        txt = tr.get_text(strip=True, separator=" ")
                        if "Issue Type" in txt and "SME" in txt.upper():
                            stock_type = "SME"
                        if "Listing At" in txt:
                            listing_txt = txt.upper()
                            if "SME" in listing_txt:
                                stock_type = "SME"
                            if "NSE" in listing_txt and "BSE" in listing_txt:
                                exchange = "NSE / BSE"
                            elif "BSE" in listing_txt:
                                exchange = "BSE"
                            elif "NSE" in listing_txt:
                                exchange = "NSE"
                        if issue_price == 0.0 and any(k in txt for k in ["Issue Price", "Final Issue Price", "Price Band"]):
                            nums = re.findall(r"\d+(?:\.\d+)?", txt)
                            valid_nums = [float(n) for n in nums if 0 < float(n) < 100000]
                            if valid_nums:
                                issue_price = max(valid_nums)
                        if lot_size == 1 and "Lot Size" in txt:
                            nums = re.findall(r"[\d,]+", txt)
                            clean_nums = [int(n.replace(",", "")) for n in nums if n.replace(",", "").isdigit()]
                            if clean_nums:
                                lot_size = clean_nums[0]
                    break
            except Exception:
                pass
    except ImportError:
        pass

    # 3. Fallback: Yahoo Finance API for Exchange lookup
    if exchange == "NSE":
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            r = requests.get(f"https://query2.finance.yahoo.com/v1/finance/search?q={symbol_clean}", headers=headers, timeout=5)
            if r.status_code == 200:
                quotes = r.json().get("quotes", [])
                exchanges_found = [q.get("exchange") for q in quotes if q.get("exchange")]
                if any(ex in ["BSE"] for ex in exchanges_found) and not any(ex in ["NSI"] for ex in exchanges_found):
                    exchange = "BSE"
        except Exception:
            pass

    if lot_size >= 100:
        stock_type = "SME"

    return stock_type, exchange, issue_price, lot_size


def get_stock_type(symbol_or_name: str, lot_size: int = 1) -> str:
    stock_type, _, _, _ = get_stock_full_details(symbol_or_name, existing_lot_size=lot_size)
    return stock_type


def get_exchange(symbol_or_name: str, lot_size: int = 1) -> str:
    _, exchange, _, _ = get_stock_full_details(symbol_or_name, existing_lot_size=lot_size)
    return exchange


def get_issue_price(symbol_or_name: str) -> float:
    _, _, issue_price, _ = get_stock_full_details(symbol_or_name)
    return issue_price


def get_lot_size(symbol_or_name: str) -> int:
    _, _, _, lot_size = get_stock_full_details(symbol_or_name)
    return lot_size


def excel_holdings(usr_id: str, holding_symbol: str = None, shares_allocated: int = None):
    """
    Scans the user's sheet in allotted_holdings.xlsx for rows where
    special_session_status == 5 (pending enrichment)
    and fills ONLY:
      - Col 1: security_name (symbol)
      - Col 2: lot_size
      - Col 3: issue_price
      - Col 4: shares_allocated (number of shares)
      - Col 5: lots_issued
      - Col 6: exchange (NSE / BSE)
      - Col 7: stock_category (Mainboard / SME)
    """
    print("excel_holdings", usr_id)
    path = get_allotted_holdings_path()
    if not os.path.exists(path):
        print(f"excel_holdings Error: File not found at '{path}'")
        return

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"excel_holdings Error opening workbook '{path}': {e}")
        return

    sheet_name = str(usr_id)
    if sheet_name not in wb.sheetnames:
        print(f"excel_holdings Error: Sheet '{sheet_name}' not found in '{path}'. Available sheets: {wb.sheetnames}")
        wb.close()
        return

    ws = wb[sheet_name]

    # If a holding_symbol was detected, check if it's already in the sheet
    if holding_symbol:
        holding_symbol = str(holding_symbol).strip()
        existing_symbols = [
            str(ws.cell(r, 1).value).strip().upper() 
            for r in range(2, ws.max_row + 1) 
            if ws.cell(r, 1).value
        ]
        if holding_symbol.upper() not in existing_symbols:
            next_row = ws.max_row + 1
            ws.cell(next_row, 1, holding_symbol)
            if shares_allocated is not None:
                ws.cell(next_row, 4, shares_allocated)
            ws.cell(next_row, 8, 5)
            try:
                wb.save(path)
                print(f"excel_holdings: Appended new holding '{holding_symbol}' (Shares={shares_allocated}) at row {next_row}")
                try:
                    from common_foundation import send_telegram_notification
                    send_telegram_notification(
                        f"🎉 <b>New IPO Allotment Detected!</b>\n"
                        f"<b>Security</b>: {holding_symbol}\n"
                        f"<b>Account (UCI)</b>: {usr_id}\n"
                        f"<b>Shares Allocated</b>: {shares_allocated or 'N/A'}"
                    )
                except Exception:
                    pass
            except Exception as e:
                print(f"excel_holdings Error saving new holding row: {e}")
        else:
            # Update shares_allocated for existing row if provided
            for r in range(2, ws.max_row + 1):
                cell_val = str(ws.cell(r, 1).value or "").strip()
                if cell_val.upper() == holding_symbol.upper():
                    if shares_allocated is not None:
                        ws.cell(r, 4, shares_allocated)
                    break

    max_r = max(ws.max_row, 20)
    for k in range(2, max_r + 1):
        status_cell = ws.cell(k, 8).value if ws.cell(k, 8).value is not None else ws.cell(k, 7).value
        if status_cell is None:
            continue

        try:
            spl_session_status = int(status_cell)
        except (ValueError, TypeError):
            continue

        if spl_session_status != 5:
            continue

        name1 = ws.cell(k, 1).value
        if name1 is None or str(name1).strip() == "" or str(name1).strip() == "None":
            continue

        name1 = str(name1).strip()
        existing_lot = ws.cell(k, 2).value or 1
        try:
            existing_lot = int(existing_lot)
        except (ValueError, TypeError):
            existing_lot = 1

        existing_price = ws.cell(k, 3).value or 0.0
        try:
            existing_price = float(existing_price)
        except (ValueError, TypeError):
            existing_price = 0.0

        existing_shares = ws.cell(k, 4).value or 0
        try:
            existing_shares = int(existing_shares)
        except (ValueError, TypeError):
            existing_shares = 0

        if shares_allocated is not None and holding_symbol and holding_symbol.upper() == name1.upper():
            existing_shares = shares_allocated

        print(f"Processing row {k}: {name1}")

        # Fetch ONLY stock symbol, exchange, stock_category, issue_price, lot_size
        try:
            stock_type, exchange, issue_price, lot_size = get_stock_full_details(
                name1, existing_lot_size=existing_lot, existing_price=existing_price
            )
        except Exception as e:
            print(f"excel_holdings: details fetch failed at row {k}: {e}")
            stock_type, exchange, issue_price, lot_size = "Mainboard", "NSE", existing_price, existing_lot

        # Calculate lots issued
        lots_issued = (existing_shares // lot_size) if (lot_size > 0 and existing_shares > 0) else (ws.cell(k, 5).value or 0)

        # Update ONLY the requested columns in Excel
        ws.cell(k, 1, name1)             # Col 1: security_name (symbol)
        ws.cell(k, 2, lot_size)          # Col 2: lot_size
        ws.cell(k, 3, issue_price)       # Col 3: issue_price
        ws.cell(k, 4, existing_shares)   # Col 4: shares_allocated
        ws.cell(k, 5, lots_issued)      # Col 5: lots_issued
        ws.cell(k, 6, exchange)          # Col 6: exchange
        ws.cell(k, 7, stock_type)        # Col 7: stock_category

        try:
            wb.save(path)
            print(f"excel_holdings: Successfully updated details for row {k} ({name1}): Shares={existing_shares}, Lots={lots_issued}, LotSize={lot_size}, Price={issue_price}, Exchange={exchange}, Category={stock_type}")
        except PermissionError:
            print(f"excel_holdings Error: Permission denied. Please ensure '{path}' is closed.")
        except Exception as e:
            print(f"excel_holdings An error occurred while saving the file: {e}")

    wb.close()


def allotment_update():
    from Base import load_credentials
    users = load_credentials("credentials.json")
    for user in users:
        client_id = user.get("broker_client_id")
        password_user = user.get("password_broker")
        topt_broker = user.get("topt_broker")
        if client_id and password_user and topt_broker:
            fetch_allotment_holdings(
                user_id=client_id,
                password=password_user,
                totp_secret=topt_broker,
            )
