from Base import load_credentials
import pandas as pd
import time
from zerodha_sell import zerodha_sell
import pandas as pd
import openpyxl
from datetime import date,timedelta
import kotak_get_balance
import time
import asyncio
from zerodha_withdraw import withdraw_from_zerodha
from Base import *

CREDENTIALS_FILE = "credentials.json"

def ipo_required_fund(d=0):
    total_sme = 0
    total_mb = 0
    row_sme = get_last_row_sme() - 1
    row_mb = get_last_row_mb() - 1
    total_sme_1 = 0
    total_mb_1 = 0

    path = '../General.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    sme_ws = wb['IPOSME']
    main_ws = wb['IPOMB']
    sme_fund = 0
    mb_fund = 0
    target_date = date.today() + timedelta(days=d)
    target_day = target_date.day
    print("date:", target_day)
    for i in range(0, 9):
        rw = row_sme - i
        apply = sme_ws.cell(rw, 42).value
        # print(apply)
        close_date = sme_ws.cell(rw, 40).value

        if apply == 3:
            if target_day == close_date:
                # print("target_day is IPO at row:", rw)
                sme_fund = sme_fund + 280000
                total_sme = total_sme + 1

    print("sme fund", sme_fund)

    for i in range(0, 9):
        rw = row_mb - i
        apply = main_ws.cell(rw, 42).value
        # print(apply)
        close_date = main_ws.cell(rw, 40).value
        if apply == 3:
            if target_day == close_date:
                mb_fund = mb_fund + 209000
                total_mb = total_mb + 1

    print("mb fund", mb_fund)
    total_fund = sme_fund + mb_fund
    print("Total Fund Required on:", target_day, "is:", total_fund)
    return total_fund

def priority_ipo_sell_smws():
    print("priority_ipo_sell_smws")
    required_fund_today = ipo_required_fund(0)
    required_fund_tomorrow = ipo_required_fund(1)
    if required_fund_tomorrow != 0:
        users = load_credentials(CREDENTIALS_FILE)
        # Users
        for user in users:
            client_id = user.get("broker_client_id")
            password_user = user.get("password_broker")
            topt_broker = user.get("topt_broker")
            bank_user = user.get("bank_user")
            bank_password = user.get("bank_password")
            email_user = user.get("email_user")
            email_password = user.get("email_password")
            try:
                #balance = 120000
                balance = asyncio.run(kotak_get_balance.get_kotak_balance(USER_ID=bank_user, PASSWORD=bank_password, EMAIL_USR=email_user,EMAIL_PSS=email_password))
            except Exception as e:
                print(e)
                balance = 0
            carryover_bank_balance = balance - required_fund_today
            if carryover_bank_balance < 0:
                carryover_bank_balance = 0
            print("carryover_bank_balance", carryover_bank_balance)
            money_need_tomorrow = required_fund_tomorrow - carryover_bank_balance
            print(client_id, money_need_tomorrow)
            if money_need_tomorrow > 2000:

                zerodha_sell(user_id=client_id, password=password_user, totp_secret=topt_broker,
                             security_symbol="NIFTYIETF")

                zerodha_sell(user_id=client_id, password=password_user, totp_secret=topt_broker,
                             security_symbol="TATAGOLD")

                zerodha_sell(user_id=client_id, password=password_user, totp_secret=topt_broker,
                             security_symbol="TATSILV")
            return 1
    else:
        print("No withdrawal required.")
        return 0

#priority_ipo_sell_smws()