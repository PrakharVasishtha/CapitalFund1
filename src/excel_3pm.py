import openpyxl
import time
from Base import get_vix
from ExtractGMP import get_ipo_gmp
from ExtractReview import has_dicey_word
from ExtractSubscription import get_ipo_subscription_live
from datetime import date
from foundation import *


def update_row_3pm(row: int, type1: str):
        print("update_row_3pm", row, type1)
        path = '../General.xlsx'
        wb = openpyxl.load_workbook(path)
        sme_ws = wb['IPOSME']
        main_ws = wb['IPOMB']
        #print(f"---------update_row_3pm---------Row: {row}")
        ws = sme_ws if type1 == "SME" else main_ws
        time.sleep(.2)
        
        
        today = date.today().day
        close_day = ws.cell(row, 40).value
        if today == close_day or today + 1 == close_day:
            url2 = ws.cell(row, 1).value
            name1 = ws.cell(row, 2).value
            print(name1)
            if name1 != None:
                #
                url1 = ws.cell(row, 3).value

                gmp = get_ipo_gmp(name1)
                print(gmp)
                sub = get_ipo_subscription_live(url2)
                print(sub)
                review = int(has_dicey_word(url2))
                AI = get_vix()
                ws.cell(row, 28, gmp)
                ws.cell(row, 23, review)
                ws.cell(row, 24, sub[0])
                ws.cell(row, 25, sub[1])
                ws.cell(row, 26, sub[2])
                ws.cell(row, 35, AI)
            else:
                print("name is empty in sheet at:",row)
            try:
                wb.save(path)
                print(f"update_3pm Successfully updated details for row {row}")
            except PermissionError:
                print(f"update_3pm Error: Permission denied. Please ensure '{path}' is closed.")
            except Exception as e:
                print(f"update_3pm An error occurred while saving the file: {e}")
        else:
            #print("Closing not today")
            x=0
        wb.close()

update_row_3pm(114,'SME')
#update_row_3pm(68,'MB')