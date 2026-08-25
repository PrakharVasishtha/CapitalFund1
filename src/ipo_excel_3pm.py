import openpyxl
import time
from Base import get_vix, get_excel_path, safe_load_workbook, safe_save_workbook
from ipo_ExtractGMP import get_ipo_gmp
from ipo_ExtractReview import has_dicey_word
from ipo_ExtractSubscription import get_ipo_subscription_live
from common_foundation import dprint


def update_row_3pm(row: int, type1: str):
    path = get_excel_path()
    wb = safe_load_workbook(path)
    try:
        sme_ws = wb['IPOSME']
        main_ws = wb['IPOMB']
        ws = sme_ws if type1 == "SME" else main_ws
        time.sleep(.1)

        try:
            raw_close = ws.cell(row, 40).value
            close_day = int(float(str(raw_close).strip())) if raw_close is not None else -1
        except Exception:
            close_day = -1

        import datetime
        today_date = datetime.date.today()
        today_day = today_date.day
        tomorrow_day = (today_date + datetime.timedelta(days=1)).day
        
        # Match today, tomorrow, or active closing day
        is_closing_target = (close_day > 0) and (close_day == today_day or close_day == tomorrow_day or abs(close_day - today_day) <= 2)

        if is_closing_target:
            url2 = ws.cell(row, 1).value
            name1 = ws.cell(row, 2).value
            dprint(name1)
            if name1 is not None:
                gmp = get_ipo_gmp(name1)
                dprint(f"GMP for {name1}: {gmp}")
                sub = get_ipo_subscription_live(url2)
                dprint(f"Sub for {name1}: {sub}")
                review = int(has_dicey_word(url2))
                AI = get_vix()
                ws.cell(row, 28, gmp)
                ws.cell(row, 23, review)
                if isinstance(sub, list) and len(sub) >= 3:
                    ws.cell(row, 24, sub[0])
                    ws.cell(row, 25, sub[1])
                    ws.cell(row, 26, sub[2])
                ws.cell(row, 35, AI)

                # Refresh Anchor Allocation (Col 22)
                try:
                    from IpoDataExtractor import ChittorgarhIPOExtractor
                    extractor = ChittorgarhIPOExtractor()
                    data_ext = extractor.extract(url2)
                    ws.cell(row, 22, int(data_ext.anchor_allocation))
                except Exception as a_err:
                    pass

                # Recalculate formula scores directly in memory
                try:
                    from ipo_formula import Formula
                    fr = Formula()
                    nw = float(ws.cell(row, 6).value or 0)
                    tb = float(ws.cell(row, 5).value or 0)
                    ti = float(ws.cell(row, 15).value or 0)
                    pat = float(ws.cell(row, 9).value or 0)
                    sqib = float(ws.cell(row, 24).value or 0)
                    sni = float(ws.cell(row, 25).value or 0)
                    gmp_val = float(ws.cell(row, 28).value or 0)
                    peaftr = float(ws.cell(row, 18).value or 0)
                    tiyoy = float(ws.cell(row, 31).value or 0)

                    G = round((100 * (nw - tb) / nw), 2) if nw != 0 else 0
                    P = round((ti / pat), 2) if pat != 0 else 0
                    ws.cell(row, 7, G)
                    ws.cell(row, 16, P)

                    if type1 == "SME":
                        anchor = float(ws.cell(row, 22).value or 0)
                        AA = fr.snr_sme(sqib, sni)
                        AC = fr.gmp_effect_sme(gmp_val)
                        AG = fr.pre_sub_aggregate_sme(G, P, peaftr, anchor, tiyoy)
                        AH = fr.post_sub_aggregate_sme(AG, sqib, sni, AA, AC)
                        ws.cell(row, 27, AA)
                        ws.cell(row, 29, AC)
                        ws.cell(row, 33, AG)
                        ws.cell(row, 34, AH)
                    else:
                        sri = float(ws.cell(row, 26).value or 0)
                        pryoy = float(ws.cell(row, 30).value or 0)
                        nwyoy = float(ws.cell(row, 32).value or 0)
                        AA = fr.snr_mb(sqib, sni)
                        AC = fr.gmp_effect_mb(gmp_val)
                        AG = fr.pre_sub_aggregate_mb(G, P, peaftr, pryoy, tiyoy, nwyoy)
                        AH = fr.post_sub_aggregate_mb(AG, sqib, sni, sri, AA, AC)
                        AI_tot = fr.total_mb(AG, AH)
                        ws.cell(row, 27, AA)
                        ws.cell(row, 29, AC)
                        ws.cell(row, 33, AG)
                        ws.cell(row, 34, AH)
                        ws.cell(row, 35, AI_tot)
                except Exception as f_err:
                    print(f"Formula calculation notice for row {row}: {f_err}")

                try:
                    safe_save_workbook(wb, path)
                    print(f"update_3pm Successfully updated details for row {row} ({name1})")
                except PermissionError:
                    print(f"update_3pm Error: Permission denied. Please ensure '{path}' is closed.")
                except Exception as e:
                    print(f"update_3pm An error occurred while saving the file: {e}")
            else:
                print("name is empty in sheet at:", row)
    finally:
        wb.close()

#update_row_3pm(182,'SME')
#update_row_3pm(68,'MB')