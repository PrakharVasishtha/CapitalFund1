from ipo_cleanfetcheddata import clean_ipo_data
from IpoDataExtractor import ChittorgarhIPOExtractor, IPOData
import openpyxl
import imaplib
import requests
from bs4 import BeautifulSoup
import re
import time
import datetime
from email import message_from_bytes
from email.utils import parsedate_to_datetime
from typing import Dict, Any
import yfinance as yf
import json
import os
from dotenv import load_dotenv

# Loads variables from a local .env file (gitignored) into the environment.
# Safe to call even if .env doesn't exist - it just does nothing.
load_dotenv()


def parse_float(val: Any) -> float:
    if isinstance(val, (int, float)):
        return float(val)
    if not isinstance(val, str):
        return 0.0
    val = val.replace('%', '').replace('x', '').replace(',', '').replace('₹', '').replace('Cr', '').strip()
    try:
        return float(val)
    except ValueError:
        return 0.0


def load_credentials(file_path: str = "credentials.json") -> list:
    """Load user credentials.

    Credentials now live in the CAPITALFUND_USERS environment variable
    (set via a local .env file, which is gitignored) instead of a
    plaintext JSON file in the repo. The `file_path` argument is kept
    for backward compatibility with existing call sites but is no
    longer read from disk.

    CAPITALFUND_USERS must be a JSON array of user objects, e.g.:

        CAPITALFUND_USERS='[{"uci": "1", "name": "...", ...}, {...}]'

    See .env.example for the full expected shape.
    """
    raw = os.environ.get("CAPITALFUND_USERS")
    if not raw:
        print(
            "Error: CAPITALFUND_USERS environment variable not set. "
            "Copy .env.example to .env in the project root and fill in "
            "your real credentials, or set the variable another way "
            "(CI secrets, systemd EnvironmentFile, etc.)."
        )
        return []
    try:
        users = json.loads(raw)
        if not isinstance(users, list):
            print("Error: CAPITALFUND_USERS must be a JSON array of user objects.")
            return []
        return users
    except json.JSONDecodeError as e:
        print(f"Error: CAPITALFUND_USERS is not valid JSON: {e}")
        return []


def Logger(file,StringText="OK",FunctionName="In Function"):
    s = FunctionName+str(StringText)+" at time :"+str(datetime.datetime.now()) 
    f = open(file, "a")
    f.write(s)
    f.write("\n")
    f.close()
    
def get_vix():
    try:
        vix_value = yf.Ticker("^INDIAVIX").info.get('regularMarketPrice')
        #print("Current India VIX:", round(vix_value, 2) if vix_value else "N/A")
    except:
        vix_value = 13.5
    return vix_value


def get_netbanking_otp(EMAIL_USER,EMAIL_PASS, search_query, retries=7, delay=4):
    for i in range(retries):
        print(f"Attempt {i + 1}: Searching for OTP...")
        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(EMAIL_USER, EMAIL_PASS)
            mail.select("inbox")

            # Search for the specific subject and ensure it's unread
            status, messages = mail.search(None, search_query)
            print(status, messages)
            if status == "OK" and messages[0]:
                # Get the list of matching message IDs and iterate backwards from the latest
                message_ids = messages[0].split()
                print(message_ids)
                for latest_id in reversed(message_ids):
                    _, msg_data = mail.fetch(latest_id, "(RFC822)")
                    msg = message_from_bytes(msg_data[0][1])

                    # --- NEW TIME FILTERING CODE ---
                    # Parse the 'Date' header to a datetime object
                    email_date = parsedate_to_datetime(msg.get("Date"))
                    now = datetime.datetime.now(datetime.timezone.utc)

                    # Calculate the difference in minutes
                    time_diff = (now - email_date).total_seconds() / 60

                    # If the email is older than 2 minutes, skip it
                    if time_diff > 2:
                        continue
                        # -------------------------------

                    body = ""
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        if content_type in ["text/html", "text/plain"]:
                            payload = part.get_payload(decode=True).decode(errors='ignore')
                            body += payload

                    clean_body = re.sub(r'<[^>]+>', ' ', body)
                    otp_match = re.search(r'is\s+(\d{6})', clean_body, re.IGNORECASE)

                    if otp_match:
                        print("otp found")
                        mail.store(latest_id, '+FLAGS', '\\Seen')
                        mail.logout()
                        return otp_match.group(1)

            mail.logout()
        except Exception as e:
            print(f"Error: {str(e)}")

        time.sleep(delay)
    return None

def get_netbanking_otp_sms(EMAIL_USER, EMAIL_PASS, search_query, retries=7, delay=4):

    otp_pattern = re.compile(
        r'(?:OTP:|One\s*Time\s*Password)[^\d]*(\d{6})',
        re.IGNORECASE
    )

    for i in range(retries):
        print(f"Attempt {i + 1}: Searching for OTP...")

        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(EMAIL_USER, EMAIL_PASS)
            mail.select("inbox")

            status, messages = mail.search(None, search_query)

            print(status, messages)

            if status == "OK" and messages[0]:

                message_ids = messages[0].split()

                for latest_id in reversed(message_ids):

                    _, msg_data = mail.fetch(latest_id, "(RFC822)")
                    msg = message_from_bytes(msg_data[0][1])

                    # Time filtering
                    email_date = parsedate_to_datetime(msg.get("Date"))
                    now = datetime.datetime.now(datetime.timezone.utc)

                    time_diff = (now - email_date).total_seconds() / 60

                    if time_diff > 2:
                        continue

                    body = ""

                    for part in msg.walk():

                        content_type = part.get_content_type()

                        if content_type in ["text/plain", "text/html"]:

                            payload = part.get_payload(decode=True)

                            if payload:
                                body += payload.decode(errors="ignore")

                    clean_body = re.sub(r'<[^>]+>', ' ', body)

                    print(clean_body)

                    otp_match = otp_pattern.search(clean_body)

                    if otp_match:
                        otp = otp_match.group(1)

                        print("OTP found:", otp)

                        mail.store(latest_id, '+FLAGS', '\\Seen')
                        mail.logout()

                        return otp

            mail.logout()

        except Exception as e:
            print(f"Error: {str(e)}")

        time.sleep(delay)

    return None



def get_last_row(file,sheet):
    path = file
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    column_index = 2
    last_row = 0
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=column_index).value is not None:
            last_row = row
            break
    return last_row

def get_excel_path(filename: str = "General.xlsx") -> str:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        filename,
        os.path.join("..", filename),
        os.path.join(script_dir, filename),
        os.path.join(script_dir, "..", filename),
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            return os.path.abspath(candidate)
    return os.path.abspath(os.path.join(script_dir, "..", filename))

def get_last_row_sme():
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb['IPOSME']
    column_index = 2
    last_row = 0
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=column_index).value is not None:
            last_row = row
            break
    return last_row+1

def get_last_row_mb():
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb['IPOMB']
    column_index = 2
    last_row = 0
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=column_index).value is not None:
            last_row = row
            break
    return last_row+1


def page_contains_trust(url: str) -> bool:
    try:
        headers = {
            "User-Agent": "Mozilla/5.0"
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")
        text = soup.get_text().lower()

        return "investment trust" in text

    except Exception as e:
        print(f"Error: {e}")
        return False

def is_data_available(url):
    try:
        time.sleep(1)
        extractor = ChittorgarhIPOExtractor()
        data1 = extractor.extract(url)
        #print(data1)
        data2 = clean_ipo_data(data1)
        #print(data2)
        if "N/A" in data2.get('issue_price_per_share','N/A'):
            print("issue_price_per_share not available,so no adding")
            return False
        elif page_contains_trust(url):
            print("trust, so no adding")
            return False
        else:
            return True

    except Exception as e:
        print(f"Error: {str(e)}")
        return True

#print(is_data_available("https://www.chittorgarh.com/ipo/powerica-ipo/2570/"))
#print(is_data_available("https://www.chittorgarh.com/ipo/propshare-celestia-scheme-ipo/2965/"))
#print(is_data_available("https://www.chittorgarh.com/ipo/gsp-crop-ipo/2031/"))
#print(get_last_row_sme())
#print(get_last_row('../General.xlsx','IPOSME'))0

#print(get_netbanking_otp_sms(EMAIL_USER, EMAIL_PASS, sub))
#print(otp_shoonya(EMAIL_USER,EMAIL_PASS, search_query, retries=15, delay=7))
#print(get_vix())
