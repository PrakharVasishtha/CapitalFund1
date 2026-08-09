"""
ipo_applied_manager.py
=======================
Manages reading and writing IPO application records to IPO-applied.xlsx.

Header Schema (per user sheet):
  Col 1: IPO-Name
  Col 2: Shares Applied
  Col 3: Issue price
  Col 4: Total Application amount
"""
import os
import sys
import openpyxl

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from Base import get_excel_path, parse_float
from common_foundation import log_info, log_error


def get_ipo_applied_path() -> str:
    """Resolves absolute path to IPO-applied.xlsx."""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(base_dir, "IPO-applied.xlsx")
    if os.path.exists(path):
        return path
    if os.path.exists("IPO-applied.xlsx"):
        return "IPO-applied.xlsx"
    return path


def fetch_ipo_details_from_general(ipo_name: str, type_ipo: str = "mb") -> tuple[int, float, float]:
    """
    Fetches Shares Applied, Issue Price, and Total Application Amount for an IPO from General.xlsx.
    """
    gen_path = get_excel_path()
    shares_applied = 1
    issue_price = 0.0

    if not os.path.exists(gen_path):
        return shares_applied, issue_price, 0.0

    try:
        wb = openpyxl.load_workbook(gen_path, data_only=True)
        sheets_to_check = ["IPOMB", "IPOSME"] if type_ipo.lower() in ["mb", "mainboard"] else ["IPOSME", "IPOMB"]

        for sheet_name in sheets_to_check:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                name_val = str(ws.cell(r, 2).value or "").strip()
                if not name_val:
                    continue

                if ipo_name.lower() in name_val.lower() or name_val.lower() in ipo_name.lower():
                    # Column 7 = Issue price, Column 8/9/10 = Min shares / Lot size
                    p_val = ws.cell(r, 7).value or 0.0
                    s_val = ws.cell(r, 8).value or ws.cell(r, 9).value or 1

                    try:
                        issue_price = parse_float(str(p_val))
                    except Exception:
                        issue_price = 0.0

                    try:
                        shares_applied = int(parse_float(str(s_val)))
                        if shares_applied <= 0:
                            shares_applied = 1
                    except Exception:
                        shares_applied = 1

                    break
            if issue_price > 0:
                break

        wb.close()
    except Exception as e:
        log_error(f"Error reading General.xlsx for {ipo_name}: {e}", exc=e, function_name="fetch_ipo_details_from_general")

    total_amount = round(shares_applied * issue_price, 2)
    return shares_applied, issue_price, total_amount


def record_ipo_application(uci: str, ipo_name: str, type_ipo: str = "mb") -> bool:
    """
    Records an IPO application entry into IPO-applied.xlsx under the specified user UCI sheet.
    """
    path = get_ipo_applied_path()
    shares_applied, issue_price, total_amount = fetch_ipo_details_from_general(ipo_name, type_ipo)

    try:
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        raw_uci = str(uci).strip()
        # Truncate 'user' prefix if present (e.g. 'user1' -> '1', 'user2' -> '2')
        sheet_name = raw_uci.replace("user", "").replace("User", "").strip() if raw_uci.lower().startswith("user") else raw_uci

        if raw_uci in wb.sheetnames:
            sheet_name = raw_uci

        header = ["IPO-Name", "Shares Applied", "Issue price", "Total Application amount"]

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(header)
        else:
            ws = wb[sheet_name]
            # Verify header
            if ws.max_row == 0 or ws.cell(1, 1).value != "IPO-Name":
                ws.cell(1, 1, header[0])
                ws.cell(1, 2, header[1])
                ws.cell(1, 3, header[2])
                ws.cell(1, 4, header[3])

        # Check if entry already exists
        entry_row = None
        for r in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(r, 1).value or "").strip()
            if cell_val.lower() == ipo_name.strip().lower():
                entry_row = r
                break

        if entry_row:
            ws.cell(entry_row, 2, shares_applied)
            ws.cell(entry_row, 3, issue_price)
            ws.cell(entry_row, 4, total_amount)
            log_info(f"Updated existing application entry for '{ipo_name}' in sheet '{sheet_name}' (Row {entry_row})", "record_ipo_application")
        else:
            new_row = ws.max_row + 1
            ws.cell(new_row, 1, ipo_name.strip())
            ws.cell(new_row, 2, shares_applied)
            ws.cell(new_row, 3, issue_price)
            ws.cell(new_row, 4, total_amount)
            log_info(f"Appended new application entry for '{ipo_name}' in sheet '{sheet_name}' (Row {new_row})", "record_ipo_application")

        wb.save(path)
        wb.close()
        return True

    except Exception as e:
        log_error(f"Failed to record IPO application in '{path}': {e}", exc=e, function_name="record_ipo_application")
        return False


if __name__ == "__main__":
    # Test recording
    res = record_ipo_application("test_uci", "Sample Tech IPO", "mb")
    print("Record test result:", res)
