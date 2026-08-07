from datetime import date

import openpyxl

from Base import *
from allotment_kotak_ipo_apply import apply_to_ipo, apply_to_ipo_all_users


def IPO_to_apply():
    print("--------IPO to Apply-------")
    IPO_sme_1 = []
    IPO_sme_2 = []
    IPO_sme_3 = []
    IPO_mb_1 = []
    IPO_mb_2 = []
    IPO_mb_3 = []

    row_sme = get_last_row_sme() - 1
    row_mb = get_last_row_mb() - 1


    path = get_excel_path()
    wb = openpyxl.load_workbook(path, data_only=True)
    sme_ws = wb['IPOSME']
    main_ws = wb['IPOMB']

    for i in range(0, 9):
        rw = row_sme - i
        apply = sme_ws.cell(rw, 42).value
        name = sme_ws.cell(rw, 2).value
        name = name[:15]
        if apply == 1:
            close_date = sme_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_sme_1.append(name)

        if apply == 2:
            close_date = sme_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_sme_2.append(name)
        
        if apply == 3:
            close_date = sme_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_sme_3.append(name)

    for i in range(0, 9):
        rw = row_mb - i
        apply = main_ws.cell(rw, 42).value
        name = main_ws.cell(rw, 2).value
        name = name[:15]
        if apply == 1:
            close_date = main_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_mb_1.append(name)

        if apply == 2:
            close_date = main_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_mb_2.append(name)
        
        if apply == 3:
            close_date = main_ws.cell(rw, 40).value
            today = date.today().day
            if today == close_date:
                IPO_mb_3.append(name)
    return IPO_mb_3,IPO_sme_3,IPO_mb_2, IPO_sme_2, IPO_mb_1, IPO_sme_1,

def apply_ipo_sme(ipo):
    print("Appling IPO:", ipo)

def apply_ipo_mb(ipo):
    print("Appling IPO:", ipo)
    #mb first try in snii category if not sufficient funds, it will try in retail.

def ipo_application():
    print(
        "*************************************-----------ipo_application----------************************************")
    print(
        "##############################################################################################################")
    all=IPO_to_apply()
    #print(all)
    IPO_mb_3 = all[0]
    IPO_sme_3 = all[1]
    IPO_mb_2 = all[2]
    IPO_sme_2 = all[3]
    IPO_sme_1 = all[4]
    IPO_mb_1 = all[5]

    print("applying for ipos:",all)
    for ipo in IPO_mb_3:
        print(ipo)
        try:
            apply_to_ipo_all_users(ipo_name = ipo,type_ipo = "mb")
        except Exception as e:
            print(e)


    for ipo in IPO_sme_3:
        print(ipo)
        try:
            apply_to_ipo_all_users(ipo_name = ipo,type_ipo = "sme")
        except Exception as e:
            print(e)
    
    for ipo in IPO_mb_2:
        print(ipo)
        try:
            apply_to_ipo_all_users(ipo_name = ipo,type_ipo = "mb")
        except Exception as e:
            print(e)
        
    for ipo in IPO_sme_2:
        print(ipo)
        try:
            apply_to_ipo_all_users(ipo_name = ipo,type_ipo = "sme")
        except Exception as e:
            print(e)

    for ipo in IPO_mb_1:
        print(ipo)
        try:
            apply_to_ipo_all_users(ipo_name = ipo,type_ipo = "mb")
        except Exception as e:
            print(e)

    for ipo in IPO_sme_1:
        print(ipo)
        try:
            apply_to_ipo_all_users(ipo_name = ipo,type_ipo = "sme")
        except Exception as e:
            print(e)

#print(IPO_to_apply())
#ipo_application()