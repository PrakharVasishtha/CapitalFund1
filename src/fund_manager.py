import pandas as pd
import openpyxl
from datetime import date,timedelta
import fund_kotak_get_balance
import time
import asyncio

from fund_zerodha_withdraw import withdraw_from_zerodha
from Base import get_last_row_sme, get_last_row_mb, get_excel_path, load_credentials

CREDENTIALS_FILE = "credentials.json"


def strategy_status_10days():
    #row_sme = get_last_row_sme() - 1

    x = "Loading..."
    i = 0
    df = None
    while (x == "Loading..." or x == "nan") and i < 10:
        url_csv = f"https://docs.google.com/spreadsheets/d/e/2PACX-1vSs2i_IJgQNpj8_gd4OMMQvvMh-G2iO15FPlMm-x3Z8lYTjX0-BePODzuXzTKq-bFZZHmyqCueCtx-5/pub?output=csv&t={int(time.time())}"
        try:
            df = pd.read_csv(url_csv)
            x = str(df.iloc[23, 4]).strip()
        except Exception as e:
            print("Error reading Google Sheet CSV:", e)
        time.sleep(2)
        i = i + 1
    if i == 10 or x == "Loading..." or df is None:
        print("sheet not loading")
        return 0
        
    try:
        buynifty = df.iloc[23, 4]
        goldetfbuy = df.iloc[26, 4]
        silveretfbuy = df.iloc[29, 4]
        return int(buynifty)+int(goldetfbuy)+int(silveretfbuy)
    except (ValueError, TypeError) as e:
        print("Invalid signal value in sheet:", e)
        return 0

def ipo_required_fund(d = 0):
    row_sme = get_last_row_sme() - 1
    row_mb = get_last_row_mb() - 1
    total_sme_1 = 0
    total_sme_2 = 0
    total_mb_1 = 0
    total_mb_2 = 0

    path = get_excel_path()
    wb = openpyxl.load_workbook(path, data_only=True)
    sme_ws = wb['IPOSME']
    main_ws = wb['IPOMB']
    sme_fund = 0
    mb_fund = 0
    buy=int(strategy_status_10days())
    print("buy:",buy)
    target_date = date.today() + timedelta(days=d)
    target_day = target_date.day
    print("date:",target_day)
    for i in range(0, 9):
        rw = row_sme - i
        apply = sme_ws.cell(rw, 42).value
        #print(apply)
        close_date = sme_ws.cell(rw, 40).value

        if apply == 2 or apply == 3:
            
            if target_day == close_date:
                #print("today is IPO at row:", rw)
                sme_fund = sme_fund + 280000
                total_sme_1 = total_sme_1 + 1
                
        elif apply == 1:
            if target_day == close_date:
                if buy > 0:
                    sme_fund = sme_fund
                elif buy == 0:
                    sme_fund = sme_fund + 280000
                    total_sme_2 = total_sme_2 + 1
                
                
    print("sme fund",sme_fund)
    #print("total_sme_1:", total_sme_1)
    #print("total_sme_2:", total_sme_2)

    for i in range(0, 9):
        rw = row_mb - i
        apply = main_ws.cell(rw, 42).value
        #print(apply)
        close_date = main_ws.cell(rw, 40).value
        if apply == 2 or apply == 3:
            if target_day == close_date:
                mb_fund = mb_fund + 209000
                total_mb_1 = total_mb_1 + 1
                
        elif apply == 1:
            if target_day == close_date:
                if buy > 0:
                    mb_fund = mb_fund
                elif buy == 0:
                    mb_fund = mb_fund + 209000
                    total_mb_2 = total_mb_2 + 1


    print("mb fund", mb_fund)
    #print("total_mb_1:", total_mb_1)
    #print("total_mb_2:", total_mb_2)
    total_fund=sme_fund + mb_fund
    print("Total Fund Required on:",target_day,"is:",total_fund)
    return total_fund


def daily_money_withdraw():
    print(
        "*************************************-----------daily_money_withdraw----------************************************")
    print(
        "##############################################################################################################")
    d0 = ipo_required_fund(0)
    d1=ipo_required_fund(1)
    required_fund = d0 + (0.9*d1)
    if required_fund != 0:
        users = load_credentials(CREDENTIALS_FILE)
        #Users
        for user in users:
            uci_user = user.get("uci")
            client_id = user.get("broker_client_id")
            password_user = user.get("password_broker")
            topt_broker = user.get("topt_broker")
            bank_user = user.get("bank_user")
            bank_password = user.get("bank_password")
            email_user = user.get("email_user")
            email_password = user.get("email_password")
            try:
                #balance = 10
                balance = asyncio.run(fund_kotak_get_balance.get_kotak_balance(USER_ID=bank_user,PASSWORD=bank_password,EMAIL_USR=email_user,EMAIL_PSS= email_password))
            except Exception as e:
                print(e)
                balance = 0
            final_amount = required_fund - balance
            print("final_required_amount",final_amount)
            success, message = withdraw_from_zerodha(
                user_uci=uci_user,
                user_id=client_id,
                password=password_user,
                totp_secret=topt_broker,
                amount=final_amount,
                headless=True,
            )

            print(success, message)
            
    else:
        print("No withdrawal required.")

#print(strategy_status_10days())
#print("ipo_required_fund",ipo_required_fund(0))
#print(daily_money_withdraw())