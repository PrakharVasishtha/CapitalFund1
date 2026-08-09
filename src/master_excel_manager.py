"""
master_excel_manager.py
========================
Centralized manager for reading, updating, and synchronizing Master.xlsx.

Master.xlsx Schema ('Users' sheet):
  Col 1 : uci (Numeric ID: 1, 2)
  Col 2 : first_name
  Col 3 : last_name
  Col 4 : mobile
  Col 5 : communication email
  Col 6 : account_email
  Col 7 : intraday (Mode: 1 or 0)
  Col 8 : zerodha_access_token
  Col 9 : current_value (Portfolio valuation float)
"""
import os
import sys
import openpyxl

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from Base import load_credentials, parse_float
from common_foundation import log_info, log_error


def get_master_excel_path() -> str:
    """Resolves absolute path to Master.xlsx safely."""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(base_dir, "Master.xlsx")
    if os.path.exists(path):
        return path
    if os.path.exists("Master.xlsx"):
        return "Master.xlsx"
    return path


def sync_master_with_credentials() -> bool:
    """
    Synchronizes user accounts from credentials.json / .env (load_credentials())
    into Master.xlsx ('Users' sheet).
    """
    path = get_master_excel_path()
    users = load_credentials()

    if not users:
        log_info("sync_master_with_credentials: No user credentials found in environment.", "sync_master_with_credentials")
        return False

    try:
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()

        if "Users" not in wb.sheetnames:
            ws = wb.create_sheet("Users", 0)
        else:
            ws = wb["Users"]

        headers = [
            "uci", "first_name", "last_name", "mobile",
            "communication email", "account_email", "intraday",
            "zerodha_access_token", "current_value"
        ]

        # Ensure header row exists
        if ws.max_row == 0 or ws.cell(1, 1).value != "uci":
            for col_idx, h in enumerate(headers, 1):
                ws.cell(1, col_idx, h)

        # Get existing UCI list
        existing_ucis = {}
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val is not None:
                existing_ucis[str(val).strip()] = r

        for idx, u in enumerate(users, 1):
            raw_uci = str(u.get("uci", idx)).replace("user", "").strip()
            name_parts = str(u.get("name", "")).strip().split(" ", 1)
            first_name = name_parts[0] if len(name_parts) > 0 else ""
            last_name = name_parts[1] if len(name_parts) > 1 else ""

            account_email = u.get("email_user", "")
            intraday = u.get("intraday", "0")

            if raw_uci in existing_ucis:
                row_idx = existing_ucis[raw_uci]
            else:
                row_idx = ws.max_row + 1
                ws.cell(row_idx, 1, int(raw_uci) if raw_uci.isdigit() else raw_uci)
                existing_ucis[raw_uci] = row_idx

            if first_name and not ws.cell(row_idx, 2).value:
                ws.cell(row_idx, 2, first_name)
            if last_name and not ws.cell(row_idx, 3).value:
                ws.cell(row_idx, 3, last_name)
            if account_email and not ws.cell(row_idx, 6).value:
                ws.cell(row_idx, 6, account_email)
            if intraday is not None:
                ws.cell(row_idx, 7, str(intraday))

        wb.save(path)
        wb.close()
        log_info(f"Master.xlsx synced successfully with {len(users)} user profiles.", "sync_master_with_credentials")
        return True

    except Exception as e:
        log_error(f"Failed to sync Master.xlsx: {e}", exc=e, function_name="sync_master_with_credentials")
        return False


def update_master_user(
    uci: str,
    current_value: float = None,
    zerodha_access_token: str = None,
    amount_needed: float = None,
    intraday: str = None
) -> bool:
    """
    Updates user specific fields (current_value, zerodha_access_token, intraday, etc.)
    for a given UCI in Master.xlsx.
    """
    path = get_master_excel_path()
    if not os.path.exists(path):
        sync_master_with_credentials()

    raw_uci = str(uci).replace("user", "").strip()

    try:
        wb = openpyxl.load_workbook(path)
        if "Users" not in wb.sheetnames:
            wb.close()
            sync_master_with_credentials()
            wb = openpyxl.load_workbook(path)

        ws = wb["Users"]
        target_row = None

        for r in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(r, 1).value or "").strip()
            if cell_val == raw_uci:
                target_row = r
                break

        if not target_row:
            target_row = ws.max_row + 1
            ws.cell(target_row, 1, int(raw_uci) if raw_uci.isdigit() else raw_uci)

        if current_value is not None:
            ws.cell(target_row, 9, float(current_value))
        if zerodha_access_token is not None:
            ws.cell(target_row, 8, str(zerodha_access_token))
        if intraday is not None:
            ws.cell(target_row, 7, str(intraday))
        if amount_needed is not None:
            # Col 15 or custom column for fund withdrawal requirement
            ws.cell(target_row, 15, float(amount_needed))

        wb.save(path)
        wb.close()
        log_info(f"Updated Master.xlsx for UCI '{raw_uci}' (CurrentVal: {current_value}, Token: {'Set' if zerodha_access_token else 'N/A'})", "update_master_user")
        return True

    except Exception as e:
        log_error(f"Failed to update Master.xlsx for UCI '{uci}': {e}", exc=e, function_name="update_master_user")
        return False


if __name__ == "__main__":
    sync_master_with_credentials()
    update_master_user("1", current_value=250000.0)
    update_master_user("2", current_value=180000.0)
