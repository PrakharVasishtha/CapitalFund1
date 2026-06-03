import openpyxl
import difflib
from typing import Dict, Any
import time
import datetime
import ipo_pe
from ipo_formula import Formula
from Base import get_vix
from common_foundation import *


def safe_get(d: Any, *keys: str, default: Any = None) -> Any:
    for key in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(key, default)
    return d

def parse_float(val: Any) -> float:
    if isinstance(val, (int, float)):
        return float(val)
    if not isinstance(val, str):
        return 0.0
    val = val.replace('%', '').replace('x', '').replace(',', '').replace('₹', '').replace('Cr', '').strip()
    try:
        return float(val)
    except ValueError:
        return 0.0

def normalize_name(n: str) -> str:
    n = n.lower()
    for char in ['.', ',', '...']:
        n = n.replace(char, '')
    words = n.split()
    remove_words = {'ltd', 'ipo', 'o', 'limited', 'inc', 'corp'}  # Added more common suffixes
    words = [w for w in words if w not in remove_words]
    return ' '.join(words)

def find_boa(closing):
    date_str = closing
    date_obj = datetime.datetime.strptime(date_str, '%d%m%Y')
    dow = date_obj.strftime('%A')
    day = int(date_obj.strftime('%d'))
    print(dow,day)
    if dow == "Friday":
        print("Friday")
        boa = day + 3
    else:
        boa = day + 1
    return boa

def find_listing(closing):
    date_str = closing
    date_obj = datetime.datetime.strptime(date_str, '%d%m%Y')
    dow = date_obj.strftime('%A')
    day = int(date_obj.strftime('%d'))
    print(dow,day)
    if dow == "Wednesday":
        print("Wednesday")
        listing = day + 5
    elif dow == "Thursday":
        print("Thursday")
        listing = day + 5
    elif dow == "Friday":
        print("Thursday")
        listing = day + 5  
    else:
        listing = day + 3
    
    if listing > 30:
        listing = 1
    return listing

class ExcelManager:
    def __init__(self):
        self.path = '../General.xlsx'
        self.wb = openpyxl.load_workbook(self.path)
        self.sme_ws = self.wb['IPOSME']
        self.main_ws = self.wb['IPOMB']


    def exists(self, type1="SME", name="NONE"):
        #print("---checking if already exists---",name)
        ws = self.sme_ws if type1 == "SME" else self.main_ws
        scrape_norm = normalize_name(name)
        for cell in ws['B']:
            if cell.value:
                excel_norm = normalize_name(cell.value)
                starts1 = scrape_norm.startswith(excel_norm)
                starts2 = excel_norm.startswith(scrape_norm)
                #print(difflib.SequenceMatcher(None, excel_norm, scrape_norm).ratio())
                sim = difflib.SequenceMatcher(None, excel_norm, scrape_norm).ratio() > 0.9
                if starts1 or starts2 or sim:
                    #print("already exists")
                    return True
        return False

    def append(self, ipo: Dict,row: int):
        type_ = ipo['category']
        ws = self.sme_ws if type_ == "SME" else self.main_ws
        ws.cell(row, 2, ipo['name'])
        ws.cell(row, 3, ipo['year'])  # Adjust if column 3 is different
        self.wb.save(self.path)
        print("Appended new IPO")
        return row

    def write_details(self, row: int, ipo_data: Dict, type1: str):
        dprint("write_details")
        ws = self.sme_ws if type1 == "SME" else self.main_ws
        time.sleep(.2)
        try:
            A = safe_get(ipo_data, "url", default="")
            B = safe_get(ipo_data, "company_name", default="")
            C = safe_get(ipo_data, "year", default="2026")  # Fallback; add 'year' to clean_ipo_data if needed
            #D
            E = parse_float(safe_get(ipo_data, "financials", "total_borrowings", "31 MAR 2025", default=0))
            F = parse_float(safe_get(ipo_data, "financials", "net_worth", "31 MAR 2025", default=0))
            G = 100 * (F - E) / F if F != 0 else 0
            G = round(G, 2)
            H = parse_float(safe_get(ipo_data, "financials", "assets", "31 MAR 2025", default=0))
            I = parse_float(safe_get(ipo_data, "financials", "profit_after_tax", "31 MAR 2025", default=0))
            J = parse_float(safe_get(ipo_data, "ratios", "pat_margin", "Latest", default=0))
            K = parse_float(safe_get(ipo_data, "ratios", "eps", "Pre", default=0))
            L = parse_float(safe_get(ipo_data, "ratios", "roe", "Latest", default=0))
            M = parse_float(safe_get(ipo_data, "financials", "ebitda", "31 MAR 2025", default=0))
            N = parse_float(safe_get(ipo_data, "ratios", "ebitda_margin", "Latest", default=0))
            O = parse_float(safe_get(ipo_data, "financials", "total_income", "31 MAR 2025", default=0))
            P = (O / I) if I != 0 else 0
            P = round(P, 2)
            pe_str = safe_get(ipo_data, "ratios", "pe_ratio", default="0")
            pe_values = [parse_float(v) for v in pe_str.split(',') if v.strip()]
            Q = pe_values[0] if pe_values else 0.0
            R = pe_values[1] if len(pe_values) > 1 else 0.0
            try:
                S = pe.effect_pe_sme(A) if type1 == "SME" else pe.effect_pe_mb(A)
            except Exception as e:
                print(e)
                S = 0
            T = parse_float(safe_get(ipo_data, "ipo_details", "issue_size", default="1"))
            U = parse_float(safe_get(ipo_data, "ipo_details", "offer_for_sale", default="1"))
            V = int(safe_get(ipo_data, "anchor_allocation", default="0"))


            AD25 = parse_float(safe_get(ipo_data, "financials", "profit_after_tax", "31 MAR 2025", default=0))
            AD24 = parse_float(safe_get(ipo_data, "financials", "profit_after_tax", "31 MAR 2024", default=0))
            AD = ((AD25 - AD24) / AD24 * 100) if AD24 != 0 else 0
            AD = round(AD, 2)
            AE25 = parse_float(safe_get(ipo_data, "financials", "total_income", "31 MAR 2025", default=0))
            AE24 = parse_float(safe_get(ipo_data, "financials", "total_income", "31 MAR 2024", default=0))
            AE = ((AE25 - AE24) / AE24 * 100) if AE24 != 0 else 0
            AE = round(AE, 2)
            AF25 = parse_float(safe_get(ipo_data, "financials", "net_worth", "31 MAR 2025", default=0))
            AF24 = parse_float(safe_get(ipo_data, "financials", "net_worth", "31 MAR 2024", default=0))
            AF = ((AF25 - AF24) / AF24 * 100) if AF24 != 0 else 0
            AF = round(AF, 2)
            AK = B
            AL = safe_get(ipo_data, "ipo_timeline", "close", default="35112026")
            print("date",AL)
            closing = int(AL[:2])
            boa = find_boa(AL)
            listing = find_listing(AL)
            
            
        except Exception as e:
            print(f"write_details An error occurred while assigning data for row {row}: {e}")
            return  # Skip save on error

        # Write to cells (using cell(row, col) for reliability)
        ws.cell(row, 1, A)   # A
        ws.cell(row, 2, B)   # B
        ws.cell(row, 3, C)     #C

        ws.cell(row, 5, E)   # E
        ws.cell(row, 6, F)   # F
        ws.cell(row, 7, G)   # G
        ws.cell(row, 8, H)   # H
        ws.cell(row, 9, I)   # I
        ws.cell(row, 10, J)  # J PAT% (10)
        ws.cell(row, 11, K)  # K
        ws.cell(row, 12, L)  # L
        ws.cell(row, 13, M)  # M
        ws.cell(row, 14, N)  # N
        ws.cell(row, 15, O)  # O
        ws.cell(row, 16, P)  # P
        ws.cell(row, 17, Q)  # Q
        ws.cell(row, 18, R)  # R
        ws.cell(row, 19, S)  # R
        ws.cell(row, 20, T)  # T Total Issue Size (20)
        ws.cell(row, 21, U)  # U
        ws.cell(row, 22, V)  #V

        ws.cell(row, 30, AD) # AD (30)
        ws.cell(row, 31, AE) # AE
        ws.cell(row, 32, AF) #AF

        ws.cell(row, 39, AK) # AK (37)
        ws.cell(row, 40, closing) # AL
        ws.cell(row, 41, boa) #AM
        #ws.cell(row, 42, listing) #AM
        
        try:
            self.wb.save(self.path)
            print(f"write_details: Successfully updated details {B}for row {row}")
        except PermissionError:
            print(f"write_details Error: Permission denied. Please ensure '{self.path}' is closed.")
        except Exception as e:
            print(f"write_details An error occurred while saving the file: {e}")
            
    def write_details_GSR(self, row: int,gmp: float, sub: list, review: int, type1: str,industry_score: float):
        dprint("write_details_GSR")
        ws = self.sme_ws if type1 == "SME" else self.main_ws
        time.sleep(.5)
        try:
            W = review
            X = parse_float(sub[0])
            Y = parse_float(sub[1])
            Z = parse_float(sub[2])
            #TS=sub[3]
            time.sleep(.5)
            AB = gmp
            time.sleep(.5)
            AQ = industry_score
            AI = get_vix()
        except Exception as e:
            print(f"write_details_GSR An error occurred while assigning data for row {row}: {e}")
            return  # Skip save on error
        time.sleep(.5)
        # Write to cells (using cell(row, col) for reliability)
        ws.cell(row, 23, W)   # A
        ws.cell(row, 24, X)   # B
        ws.cell(row, 25, Y)     #C
        ws.cell(row, 26, Z)     #C
        ws.cell(row, 28, AB)
        ws.cell(row, 35, AI)
        ws.cell(row, 43, AQ)

        try:
            self.wb.save(self.path)
            print(f"write_details_GSR: Successfully updated details for row {row}")
        except PermissionError:
            print(f"write_details_GSR Error: Permission denied. Please ensure '{self.path}' is closed.")
        except Exception as e:
            print(f"write_details_GSR An error occurred while saving the file: {e}")

    def write_formula_sme(self, row: int, type1: str):
        ws = self.sme_ws if type1 == "SME" else self.main_ws
        time.sleep(.2)
        nw = ws.cell(row, 6).value
        tb = ws.cell(row, 5).value
        ti = ws.cell(row, 15).value
        pat = ws.cell(row, 9).value
        sqib = ws.cell(row, 24).value
        sni = ws.cell(row, 25).value
        gmp = ws.cell(row, 28).value
        pebfr= ws.cell(row, 17).value
        peaftr = ws.cell(row, 18).value
        anchor = ws.cell(row, 22).value
        review = ws.cell(row, 23).value
        tiyoy = ws.cell(row, 31).value
        fr = Formula()
        try:
            G = (100 * (nw-tb)/nw)
            G = round(G, 2)
            P = ti/pat
            P = round(P, 2)
            AA= fr.snr_sme(sqib,sni)
            AC = fr.gmp_effect_sme(gmp)
            AG = fr.pre_sub_aggregate_sme(G, P, peaftr, anchor, tiyoy)
            AH = fr.post_sub_aggregate_sme(AG, sqib, sni, AA, AC)
            AI=  fr.total_sme(AG,AH)
        except Exception as e:
            print(f"write_formula_sme An error occurred while assigning data for row {row}: {e}")
            return  # Skip save on error

        # Write to cells (using cell(row, col) for reliability)
        ws.cell(row, 7, G)  # G
        ws.cell(row, 16, P)  # P
        ws.cell(row, 27, AA)  # C
        ws.cell(row, 29, AC)  # C
        ws.cell(row, 33, AG)  # C
        ws.cell(row, 34, AH)
        #ws.cell(row, 35, AI)

        try:
            self.wb.save(self.path)
            print(f"write_formula_sme Successfully updated details for row {row}")
        except PermissionError:
            print(f"write_formula_sme Error: Permission denied. Please ensure '{self.path}' is closed.")
        except Exception as e:
            print(f"write_formula_sme An error occurred while saving the file: {e}")

    def write_formula_mb(self, row: int, type1: str):
        ws = self.sme_ws if type1 == "SME" else self.main_ws
        time.sleep(.2)
        nw = 0
        tb = 0
        ti = 0
        pat = 0
        sqib = 0
        sni = 0
        sri = 0
        gmp = 0
        peaftr = 0
        pryoy = 0
        tiyoy = 0
        nwyoy = 0

        nw = ws.cell(row, 6).value
        tb = ws.cell(row, 5).value
        ti = ws.cell(row, 15).value
        pat = ws.cell(row, 9).value
        sqib = ws.cell(row, 24).value
        sni = ws.cell(row, 25).value
        sri = ws.cell(row, 26).value
        gmp = ws.cell(row, 28).value
        peaftr = ws.cell(row, 18).value
        pryoy = ws.cell(row, 30).value
        tiyoy = ws.cell(row, 31).value
        nwyoy = ws.cell(row, 32).value

        fr=Formula()

        try:
            G = (100 * (nw-tb)/nw)
            G = round(G, 2)
            P = ti/pat
            P = round(P, 2)
            AA = fr.snr_mb(sqib, sni)
            AC = fr.gmp_effect_mb(gmp)
            AG = fr.pre_sub_aggregate_mb(G, P, peaftr,pryoy, tiyoy,nwyoy)
            AH = fr.post_sub_aggregate_mb(AG, sqib, sni,sri, AA, AC)
            AI = fr.total_mb(AG, AH)

        except Exception as e:
            print(f"write_formula MB An error occurred while assigning data for row {row}: {e}")
            return  # Skip save on error

        # Write to cells (using cell(row, col) for reliability)
        ws.cell(row, 7, G)  # G
        ws.cell(row, 16, P)  # P
        ws.cell(row, 27, AA)  # C
        ws.cell(row, 29, AC)  # C
        ws.cell(row, 33, AG)  # C
        ws.cell(row, 34, AH)
        ws.cell(row, 35, AI)

        try:
            self.wb.save(self.path)
            print(f"write_formula MB Successfully updated details for row {row}")
        except PermissionError:
            print(f"write_formula MB Error: Permission denied. Please ensure '{self.path}' is closed.")
        except Exception as e:
            print(f"write_formula MB An error occurred while saving the file: {e}")


#print(ExcelManager().exists("SME","Adisoft Technologies"))
#ExcelManager().write_formula_sme(108,"SME")
#ExcelManager().write_formula_mb(69,"MB")
