import openpyxl
#import difflib
import time
#from typing import Dict, Any
#from formula import Formula
from ipo_excel_manager import parse_float
from src.ipo_formula import Formula


#from ExtractGMP import get_ipo_gmp
#from ExtractReview import has_dicey_word
#from ExtractSubscription import get_ipo_subscription_dict

def write_formula_sme(row: int, type1: str):
    path = '../General.xlsx'
    wb = openpyxl.load_workbook(path)
    sme_ws = wb['IPOSME']
    main_ws = wb['IPOMB']
    ws = sme_ws if type1 == "SME" else main_ws
    time.sleep(.2)
    nw = ws.cell(row, 6).value
    tb = ws.cell(row, 5).value
    ti = ws.cell(row, 15).value
    pat = ws.cell(row, 9).value
    sqib = ws.cell(row, 24).value
    sni = ws.cell(row, 25).value
    gmp = ws.cell(row, 28).value
    pebfr= ws.cell(row, 17).value
    peaftr = ws.cell(row, 18).value
    anchor = ws.cell(row, 22).value
    review = ws.cell(row, 23).value
    tiyoy = ws.cell(row, 31).value
    fr = Formula()
    try:
        G = (100 * (nw-tb)/nw)
        G = round(G, 2)
        P = ti/pat
        P = round(P, 2)
        AA= fr.snr_sme(sqib,sni)
        AC = fr.gmp_effect_sme(gmp)
        AG = fr.pre_sub_aggregate_sme(G, P, peaftr, anchor, tiyoy)
        AH = fr.post_sub_aggregate_sme(AG, sqib, sni, AA, AC)
        AI=  fr.total_sme(AG,AH)
    except Exception as e:
        print(f"write_formula_sme An error occurred while assigning data for row {row}: {e}")
        return  # Skip save on error

        # Write to cells (using cell(row, col) for reliability)
    ws.cell(row, 7, G)  # G
    ws.cell(row, 16, P)  # P
    ws.cell(row, 27, AA)  # C
    ws.cell(row, 29, AC)  # C
    ws.cell(row, 33, AG)  # C
    ws.cell(row, 34, AH)
    ws.cell(row, 35, AI)

    try:
        wb.save(path)
        print(f"write_formula_sme Successfully updated details for row {row}")
    except PermissionError:
        print(f"write_formula_sme Error: Permission denied. Please ensure '{path}' is closed.")
    except Exception as e:
        print(f"write_formula_sme An error occurred while saving the file: {e}")

def write_formula_mb(row: int, type1: str):
    path = '../General.xlsx'
    wb = openpyxl.load_workbook(path)
    sme_ws = wb['IPOSME']
    main_ws = wb['IPOMB']
    ws = sme_ws if type1 == "SME" else main_ws
    time.sleep(.2)
    nw = 0
    tb = 0
    ti = 0
    pat = 0
    sqib = 0
    sni = 0
    sri = 0
    gmp = 0
    peaftr = 0
    pryoy = 0
    tiyoy = 0
    nwyoy = 0

    nw = ws.cell(row, 6).value
    tb = ws.cell(row, 5).value
    ti = ws.cell(row, 15).value
    pat = ws.cell(row, 9).value
    sqib = ws.cell(row, 24).value
    sni = ws.cell(row, 25).value
    sri = ws.cell(row, 26).value
    gmp = ws.cell(row, 28).value
    peaftr = ws.cell(row, 18).value
    pryoy = ws.cell(row, 30).value
    tiyoy = ws.cell(row, 31).value
    nwyoy = ws.cell(row, 32).value

    fr=Formula()

    try:
        G = parse_float(100 * (nw-tb)/nw)

        G = round(G, 2)

        P = parse_float(ti/pat)
        P = round(P, 2)

        #AA = parse_float(fr.snr_mb(sqib, sni))
        AC = parse_float(fr.gmp_effect_mb(gmp))

        #AG = parse_float(fr.pre_sub_aggregate_mb(G, P, peaftr,pryoy, tiyoy,nwyoy))
        #AH = parse_float(fr.post_sub_aggregate_mb(AG, sqib, sni,sri, AA, AC))
        
        #AI = parse_float(fr.total_mb(AG, AH))

    except Exception as e:
        print(f"write_formula MB An error occurred while assigning data for row {row}: {e}")
        return  # Skip save on error

        # Write to cells (using cell(row, col) for reliability)
    '''
    ws.cell(row, 7, G)  # G
    ws.cell(row, 16, P)  # P
    ws.cell(row, 27, AA)  # C
    ws.cell(row, 29, AC)  # C
    ws.cell(row, 33, AG)  # C
    ws.cell(row, 34, AH)
    ws.cell(row, 35, AI)
    '''
    try:
        wb.save(path)
        print(f"write_formula MB Successfully updated details for row {row}")
    except PermissionError:
        print(f"write_formula MB Error: Permission denied. Please ensure '{path}' is closed.")
    except Exception as e:
        print(f"write_formula MB An error occurred while saving the file: {e}")


#write_formula_mb(68, 'MB')